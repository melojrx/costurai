from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Q, Count, F
from django.utils import timezone
from django.core.paginator import Paginator
from django.db import transaction
from datetime import date, datetime, timedelta
import json

from .models import (
    ContaReceber, PagamentoRecebido, FaturamentoMensal, StatusPagamento,
    CategoriaContaPagar, SubcategoriaContaPagar, ContaPagar, PagamentoEfetuado,
    ParcelaContaPagar, StatusContaPagar, TipoDocumento, FormaPagamento
)
from apps.producao.models import OrdemProducao, StatusOP
from apps.cadastros.models import Cliente
from apps.core.middleware import get_current_empresa
from decimal import Decimal


# =============================================================================
# DASHBOARD FINANCEIRO
# =============================================================================

@login_required
def dashboard_financeiro(request):
    """
    Dashboard principal do módulo financeiro
    KPIs completos de débitos e créditos
    """
    empresa = get_current_empresa()
    
    # Dados do mês atual
    hoje = timezone.now().date()
    mes_atual = hoje.month
    ano_atual = hoje.year
    inicio_mes = hoje.replace(day=1)
    fim_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    # Atualizar faturamento do mês atual
    faturamento_atual = FaturamentoMensal.atualizar_faturamento(
        empresa, mes_atual, ano_atual
    )
    
    # =====================================================================
    # KPIs PRINCIPAIS - CONTAS A RECEBER (CRÉDITOS)
    # =====================================================================
    contas_receber = ContaReceber.objects.filter(empresa=empresa)
    contas_receber_pendentes = contas_receber.filter(
        status__in=[StatusPagamento.PENDENTE, StatusPagamento.PARCIAL]
    )
    
    contas_receber_vencidas = contas_receber_pendentes.filter(data_vencimento__lt=hoje)
    contas_receber_vencendo = contas_receber_pendentes.filter(
        data_vencimento__gte=hoje,
        data_vencimento__lte=hoje + timedelta(days=7)
    )
    
    # Recebimentos do mês
    recebimentos_mes = PagamentoRecebido.objects.filter(
        empresa=empresa,
        data_pagamento__gte=inicio_mes,
        data_pagamento__lte=fim_mes
    ).aggregate(total=Sum('valor_pago'))['total'] or 0
    
    # =====================================================================
    # KPIs PRINCIPAIS - CONTAS A PAGAR (DÉBITOS)
    # =====================================================================
    contas_pagar = ContaPagar.objects.filter(empresa=empresa)
    contas_pagar_pendentes = contas_pagar.filter(
        status__in=[StatusContaPagar.PENDENTE, StatusContaPagar.PARCIAL]
    )
    
    contas_pagar_vencidas = contas_pagar_pendentes.filter(data_vencimento__lt=hoje)
    contas_pagar_vencendo = contas_pagar_pendentes.filter(
        data_vencimento__gte=hoje,
        data_vencimento__lte=hoje + timedelta(days=7)
    )
    
    # Pagamentos do mês
    pagamentos_mes = PagamentoEfetuado.objects.filter(
        empresa=empresa,
        data_pagamento__gte=inicio_mes,
        data_pagamento__lte=fim_mes
    ).aggregate(total=Sum('valor'))['total'] or 0
    
    # =====================================================================
    # FLUXO DE CAIXA - PRÓXIMOS 30 DIAS
    # =====================================================================
    data_fim_fluxo = hoje + timedelta(days=30)
    
    # Entradas previstas
    entradas_previstas = contas_receber_pendentes.filter(
        data_vencimento__gte=hoje,
        data_vencimento__lte=data_fim_fluxo
    ).aggregate(total=Sum(F('valor_total') - F('valor_recebido')))['total'] or 0
    
    # Saídas previstas
    saidas_previstas = contas_pagar_pendentes.filter(
        data_vencimento__gte=hoje,
        data_vencimento__lte=data_fim_fluxo
    ).aggregate(total=Sum(F('valor_original') + F('valor_juros') - F('valor_desconto') - F('valor_pago')))['total'] or 0
    
    # =====================================================================
    # ESTATÍSTICAS CONSOLIDADAS
    # =====================================================================
    stats = {
        # Faturamento
        'faturamento_mes': faturamento_atual,
        
        # Contas a Receber (Créditos)
        'total_a_receber': contas_receber_pendentes.aggregate(
            total=Sum(F('valor_total') - F('valor_recebido'))
        )['total'] or 0,
        'contas_receber_vencidas': {
            'count': contas_receber_vencidas.count(),
            'valor': contas_receber_vencidas.aggregate(
                total=Sum(F('valor_total') - F('valor_recebido'))
            )['total'] or 0
        },
        'contas_receber_vencendo': {
            'count': contas_receber_vencendo.count(),
            'valor': contas_receber_vencendo.aggregate(
                total=Sum(F('valor_total') - F('valor_recebido'))
            )['total'] or 0
        },
        'recebido_mes': recebimentos_mes,
        
        # Contas a Pagar (Débitos)
        'total_a_pagar': contas_pagar_pendentes.aggregate(
            total=Sum(F('valor_original') + F('valor_juros') - F('valor_desconto') - F('valor_pago'))
        )['total'] or 0,
        'contas_pagar_vencidas': {
            'count': contas_pagar_vencidas.count(),
            'valor': contas_pagar_vencidas.aggregate(
                total=Sum(F('valor_original') + F('valor_juros') - F('valor_desconto') - F('valor_pago'))
            )['total'] or 0
        },
        'contas_pagar_vencendo': {
            'count': contas_pagar_vencendo.count(),
            'valor': contas_pagar_vencendo.aggregate(
                total=Sum(F('valor_original') + F('valor_juros') - F('valor_desconto') - F('valor_pago'))
            )['total'] or 0
        },
        'pago_mes': pagamentos_mes,
        
        # Fluxo de Caixa
        'entradas_previstas': entradas_previstas,
        'saidas_previstas': saidas_previstas,
        'saldo_previsto': entradas_previstas - saidas_previstas,
        
        # Resultado do Mês
        'resultado_mes': recebimentos_mes - pagamentos_mes,
        'margem_lucro': round((recebimentos_mes - pagamentos_mes) * 100 / recebimentos_mes, 1) if recebimentos_mes > 0 else 0,
    }
    
    # =====================================================================
    # FATURAMENTO DOS ÚLTIMOS 6 MESES
    # =====================================================================
    faturamentos_recentes = []
    for i in range(6):
        data_ref = hoje.replace(day=1) - timedelta(days=i*30)
        faturamento = FaturamentoMensal.objects.filter(
            empresa=empresa,
            mes=data_ref.month,
            ano=data_ref.year
        ).first()
        
        if faturamento:
            faturamentos_recentes.append(faturamento)
        else:
            # Criar faturamento vazio para o mês
            faturamento = FaturamentoMensal.atualizar_faturamento(
                empresa, data_ref.month, data_ref.year
            )
            faturamentos_recentes.append(faturamento)
    
    faturamentos_recentes.reverse()
    
    # =====================================================================
    # CONTAS URGENTES (RECEBER E PAGAR)
    # =====================================================================
    contas_receber_urgentes = contas_receber.filter(
        status__in=[StatusPagamento.PENDENTE, StatusPagamento.PARCIAL, StatusPagamento.VENCIDO]
    ).select_related('cliente', 'ordem_producao').order_by('data_vencimento')[:5]
    
    contas_pagar_urgentes = contas_pagar.filter(
        status__in=[StatusContaPagar.PENDENTE, StatusContaPagar.PARCIAL, StatusContaPagar.VENCIDO]
    ).select_related('categoria', 'subcategoria').order_by('data_vencimento')[:5]
    
    # =====================================================================
    # DESPESAS POR CATEGORIA (TOP 5)
    # =====================================================================
    despesas_por_categoria = []
    if pagamentos_mes > 0:
        despesas_query = PagamentoEfetuado.objects.filter(
            empresa=empresa,
            data_pagamento__gte=inicio_mes,
            data_pagamento__lte=fim_mes
        ).values('conta_pagar__categoria__nome').annotate(
            total=Sum('valor')
        ).order_by('-total')[:5]
        
        for despesa in despesas_query:
            percentual = round(despesa['total'] * 100 / pagamentos_mes, 1)
            despesas_por_categoria.append({
                'categoria': despesa['conta_pagar__categoria__nome'] or 'Sem categoria',
                'total': despesa['total'],
                'percentual': percentual
            })
    else:
        despesas_por_categoria = []
    
    # =====================================================================
    # EVOLUÇÃO MENSAL (ÚLTIMOS 6 MESES)
    # =====================================================================
    evolucao_mensal = []
    for i in range(6):
        data_ref = hoje.replace(day=1) - timedelta(days=i*30)
        inicio_periodo = data_ref.replace(day=1)
        fim_periodo = (inicio_periodo + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        entradas = PagamentoRecebido.objects.filter(
            empresa=empresa,
            data_pagamento__gte=inicio_periodo,
            data_pagamento__lte=fim_periodo
        ).aggregate(total=Sum('valor_pago'))['total'] or 0
        
        saidas = PagamentoEfetuado.objects.filter(
            empresa=empresa,
            data_pagamento__gte=inicio_periodo,
            data_pagamento__lte=fim_periodo
        ).aggregate(total=Sum('valor'))['total'] or 0
        
        evolucao_mensal.append({
            'mes': data_ref.strftime('%b/%Y'),
            'entradas': entradas,
            'saidas': saidas,
            'resultado': entradas - saidas
        })
    
    evolucao_mensal.reverse()
    
    context = {
        'stats': stats,
        'faturamentos_recentes': faturamentos_recentes,
        'contas_receber_urgentes': contas_receber_urgentes,
        'contas_pagar_urgentes': contas_pagar_urgentes,
        'despesas_por_categoria': despesas_por_categoria,
        'evolucao_mensal': evolucao_mensal,
        'empresa': empresa,
        'mes_atual': mes_atual,
        'ano_atual': ano_atual,
        'hoje': hoje,
    }
    
    return render(request, 'financeiro/relatorio_financeiro.html', context)


# =============================================================================
# CONTAS A RECEBER
# =============================================================================

@login_required
def contas_receber_listar(request):
    """Lista todas as contas a receber com filtros"""
    contas = ContaReceber.objects.filter(empresa=get_current_empresa())
    
    # Filtros
    status_filter = request.GET.get('status')
    cliente_filter = request.GET.get('cliente')
    vencimento_filter = request.GET.get('vencimento')
    search = request.GET.get('search')
    
    if status_filter:
        contas = contas.filter(status=status_filter)
    if cliente_filter:
        contas = contas.filter(cliente_id=cliente_filter)
    if vencimento_filter:
        if vencimento_filter == 'vencidas':
            contas = contas.filter(data_vencimento__lt=timezone.now().date())
        elif vencimento_filter == 'vencendo':
            contas = contas.filter(
                data_vencimento__gte=timezone.now().date(),
                data_vencimento__lte=timezone.now().date() + timedelta(days=7)
            )
    if search:
        contas = contas.filter(
            Q(numero_conta__icontains=search) |
            Q(cliente__nome__icontains=search) |
            Q(ordem_producao__numero_op__icontains=search)
        )
    
    # Ordenação
    contas = contas.select_related('cliente', 'ordem_producao').order_by('-data_emissao')
    
    # Paginação
    paginator = Paginator(contas, 20)
    page = request.GET.get('page')
    contas = paginator.get_page(page)
    
    # Estatísticas
    stats = {
        'total': ContaReceber.objects.filter(empresa=get_current_empresa()).count(),
        'pendentes': ContaReceber.objects.filter(
            empresa=get_current_empresa(),
            status=StatusPagamento.PENDENTE
        ).count(),
        'vencidas': ContaReceber.objects.filter(
            empresa=get_current_empresa(),
            status=StatusPagamento.VENCIDO
        ).count(),
        'pagas': ContaReceber.objects.filter(
            empresa=get_current_empresa(),
            status=StatusPagamento.PAGO
        ).count(),
    }
    
    # Clientes para filtro
    clientes = Cliente.objects.filter(empresa=get_current_empresa()).order_by('nome')
    
    context = {
        'contas': contas,
        'stats': stats,
        'clientes': clientes,
        'status_choices': StatusPagamento.choices,
        'current_filters': {
            'status': status_filter,
            'cliente': cliente_filter,
            'vencimento': vencimento_filter,
            'search': search,
        },
        'empresa': get_current_empresa(),
    }
    
    return render(request, 'financeiro/contas_receber_listar.html', context)


@login_required
def conta_receber_detalhes(request, conta_id):
    """Detalhes de uma conta a receber com histórico de pagamentos"""
    conta = get_object_or_404(
        ContaReceber.objects.select_related('cliente', 'ordem_producao'),
        id=conta_id,
        empresa=get_current_empresa()
    )
    
    # Histórico de pagamentos
    pagamentos = conta.pagamentos.all().order_by('-data_pagamento')
    
    context = {
        'conta': conta,
        'pagamentos': pagamentos,
        'empresa': get_current_empresa(),
    }
    
    return render(request, 'financeiro/conta_receber_detalhes.html', context)


@login_required
def registrar_pagamento(request, conta_id):
    """Registrar um novo pagamento para uma conta"""
    conta = get_object_or_404(ContaReceber, id=conta_id, empresa=get_current_empresa())
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                valor_pago = float(request.POST.get('valor_pago', 0))
                data_pagamento = request.POST.get('data_pagamento')
                forma_pagamento = request.POST.get('forma_pagamento')
                observacoes = request.POST.get('observacoes', '')
                numero_documento = request.POST.get('numero_documento', '')
                
                # Validações
                if valor_pago <= 0:
                    raise ValueError("Valor deve ser maior que zero")
                
                if valor_pago > conta.valor_pendente:
                    raise ValueError("Valor não pode ser maior que o pendente")
                
                # Criar pagamento
                pagamento = PagamentoRecebido.objects.create(
                    empresa=get_current_empresa(),
                    conta_receber=conta,
                    valor_pago=valor_pago,
                    data_pagamento=data_pagamento,
                    forma_pagamento=forma_pagamento,
                    observacoes=observacoes,
                    numero_documento=numero_documento,
                    usuario_registro=request.user
                )
                
                messages.success(request, f'Pagamento de R$ {valor_pago:.2f} registrado com sucesso!')
                return redirect('financeiro:conta_detalhes', conta_id=conta.id)
                
        except Exception as e:
            messages.error(request, f'Erro ao registrar pagamento: {str(e)}')
    
    return redirect('financeiro:conta_detalhes', conta_id=conta.id)


# =============================================================================
# RELATÓRIOS DE FATURAMENTO
# =============================================================================

@login_required
def relatorio_faturamento(request):
    """
    Relatório de faturamento mensal
    Replica a aba 2 da planilha do cliente
    """
    empresa = get_current_empresa()
    
    # Filtros de período
    ano = int(request.GET.get('ano', timezone.now().year))
    mes = request.GET.get('mes')
    
    if mes:
        mes = int(mes)
        # Buscar faturamento específico
        faturamentos = [FaturamentoMensal.atualizar_faturamento(empresa, mes, ano)]
    else:
        # Buscar todos os meses do ano
        faturamentos = []
        for m in range(1, 13):
            faturamento = FaturamentoMensal.atualizar_faturamento(empresa, m, ano)
            faturamentos.append(faturamento)
    
    # Totais do ano
    totais_ano = {
        'entradas': sum(f.entradas for f in faturamentos),
        'saidas': sum(f.saidas for f in faturamentos),
        'valor_entradas': sum(f.valor_entradas for f in faturamentos),
        'valor_saidas': sum(f.valor_saidas for f in faturamentos),
        'valor_recebido': sum(f.valor_recebido for f in faturamentos),
        'falta_receber': sum(f.falta_receber for f in faturamentos),
    }
    
    context = {
        'faturamentos': faturamentos,
        'totais_ano': totais_ano,
        'ano_atual': ano,
        'mes_atual': mes,
        'empresa': empresa,
        'anos_disponiveis': range(2020, timezone.now().year + 2),
        'meses': [
            (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'),
            (4, 'Abril'), (5, 'Maio'), (6, 'Junho'),
            (7, 'Julho'), (8, 'Agosto'), (9, 'Setembro'),
            (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro')
        ]
    }
    
    return render(request, 'financeiro/relatorio_faturamento.html', context)


@login_required
def exportar_faturamento(request):
    """Exportar relatório de faturamento para Excel"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment
        
        empresa = get_current_empresa()
        ano = int(request.GET.get('ano', timezone.now().year))
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Faturamento {ano}"
        
        # Cabeçalho principal
        ws.merge_cells('A1:H1')
        ws['A1'] = f"RELATÓRIO DE FATURAMENTO - {empresa.nome} - {ano}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Cabeçalhos das colunas
        headers = [
            'Mês', 'Entradas', 'Saídas', 'A Produzir',
            'V. Entradas', 'V. Saídas', 'V. Recebido', 'Falta Receber'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Dados dos meses
        meses_nomes = [
            '', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
            'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
        ]
        
        row = 4
        for mes in range(1, 13):
            faturamento = FaturamentoMensal.atualizar_faturamento(empresa, mes, ano)
            
            ws.cell(row=row, column=1, value=meses_nomes[mes])
            ws.cell(row=row, column=2, value=faturamento.entradas)
            ws.cell(row=row, column=3, value=faturamento.saidas)
            ws.cell(row=row, column=4, value=faturamento.a_produzir)
            ws.cell(row=row, column=5, value=float(faturamento.valor_entradas))
            ws.cell(row=row, column=6, value=float(faturamento.valor_saidas))
            ws.cell(row=row, column=7, value=float(faturamento.valor_recebido))
            ws.cell(row=row, column=8, value=float(faturamento.falta_receber))
            
            row += 1
        
        # Ajustar largura das colunas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Criar response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="faturamento_{ano}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'Erro ao exportar relatório: {str(e)}')
        return redirect('financeiro:relatorio_faturamento')


# =============================================================================
# AJAX VIEWS
# =============================================================================

@login_required
def atualizar_faturamento_ajax(request):
    """Atualizar faturamento via AJAX"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            mes = data.get('mes')
            ano = data.get('ano')
            
            faturamento = FaturamentoMensal.atualizar_faturamento(
                get_current_empresa(), mes, ano
            )
            
            return JsonResponse({
                'success': True,
                'faturamento': {
                    'entradas': faturamento.entradas,
                    'saidas': faturamento.saidas,
                    'a_produzir': faturamento.a_produzir,
                    'valor_entradas': float(faturamento.valor_entradas),
                    'valor_saidas': float(faturamento.valor_saidas),
                    'valor_recebido': float(faturamento.valor_recebido),
                    'falta_receber': float(faturamento.falta_receber),
                }
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False})


@login_required
def buscar_contas_ajax(request):
    """Buscar contas a receber via AJAX"""
    query = request.GET.get('q', '')
    contas = []
    
    if query:
        contas_obj = ContaReceber.objects.filter(
            empresa=get_current_empresa(),
            numero_conta__icontains=query
        ).select_related('cliente')[:10]
        
        contas = [
            {
                'id': c.id,
                'text': f"{c.numero_conta} - {c.cliente.nome}",
                'valor_pendente': float(c.valor_pendente)
            }
            for c in contas_obj
        ]
    
    return JsonResponse({'results': contas})


@login_required
def conta_nova(request):
    """Criar nova conta a receber"""
    if request.method == 'POST':
        try:
            empresa = get_current_empresa()
            
            # Criar nova conta
            conta = ContaReceber.objects.create(
                empresa=empresa,
                cliente_id=request.POST.get('cliente'),
                ordem_producao_id=request.POST.get('ordem_producao') if request.POST.get('ordem_producao') else None,
                valor_total=Decimal(request.POST.get('valor_total')),
                data_vencimento=request.POST.get('data_vencimento'),
                observacoes=request.POST.get('observacoes', '')
            )
            
            messages.success(request, f'Conta {conta.numero_conta} criada com sucesso!')
            return redirect('financeiro:conta_detalhes', conta_id=conta.id)
            
        except Exception as e:
            messages.error(request, f'Erro ao criar conta: {str(e)}')
    
    # Buscar dados para o formulário
    empresa = get_current_empresa()
    clientes = Cliente.objects.filter(empresa=empresa).order_by('nome')
    ops = OrdemProducao.objects.filter(empresa=empresa, status__in=['planejada', 'em_producao']).order_by('-numero_op')
    
    # Buscar últimas 5 contas criadas
    ultimas_contas = ContaReceber.objects.filter(
        empresa=empresa
    ).select_related('cliente').order_by('-data_emissao')[:5]
    
    context = {
        'clientes': clientes,
        'ops': ops,
        'hoje': timezone.now().date(),
        'ultimas_contas': ultimas_contas,
    }
    
    return render(request, 'financeiro/conta_nova.html', context)


# =============================================================================
# CONTAS A PAGAR
# =============================================================================

@login_required
def contas_pagar_listar(request):
    """Lista todas as contas a pagar com filtros avançados"""
    empresa = get_current_empresa()
    contas = ContaPagar.objects.filter(empresa=empresa)
    
    # Filtros
    status_filter = request.GET.get('status')
    categoria_filter = request.GET.get('categoria')
    subcategoria_filter = request.GET.get('subcategoria')
    fornecedor_filter = request.GET.get('fornecedor')
    vencimento_de = request.GET.get('venc_de')
    vencimento_ate = request.GET.get('venc_ate')
    search = request.GET.get('search')
    
    if status_filter:
        contas = contas.filter(status=status_filter)
    if categoria_filter:
        contas = contas.filter(categoria_id=categoria_filter)
    if subcategoria_filter:
        contas = contas.filter(subcategoria_id=subcategoria_filter)
    if fornecedor_filter:
        contas = contas.filter(fornecedor_nome__icontains=fornecedor_filter)
    if vencimento_de:
        contas = contas.filter(data_vencimento__gte=vencimento_de)
    if vencimento_ate:
        contas = contas.filter(data_vencimento__lte=vencimento_ate)
    if search:
        contas = contas.filter(
            Q(numero_documento__icontains=search) |
            Q(fornecedor_nome__icontains=search) |
            Q(descricao__icontains=search)
        )
    
    # Ordenação
    contas = contas.select_related('categoria', 'subcategoria').order_by('data_vencimento', '-data_emissao')
    
    # Paginação
    paginator = Paginator(contas, 20)
    page = request.GET.get('page')
    contas = paginator.get_page(page)
    
    # Estatísticas
    hoje = timezone.now().date()
    stats = {
        'total_contas': ContaPagar.objects.filter(empresa=empresa).count(),
        'contas_vencidas': ContaPagar.objects.filter(
            empresa=empresa,
            data_vencimento__lt=hoje,
            status__in=[StatusContaPagar.PENDENTE, StatusContaPagar.PARCIAL]
        ).count(),
        'contas_vencendo': ContaPagar.objects.filter(
            empresa=empresa,
            data_vencimento__gte=hoje,
            data_vencimento__lte=hoje + timedelta(days=7),
            status__in=[StatusContaPagar.PENDENTE, StatusContaPagar.PARCIAL]
        ).count(),
        'valor_total': ContaPagar.objects.filter(
            empresa=empresa,
            status__in=[StatusContaPagar.PENDENTE, StatusContaPagar.PARCIAL]
        ).aggregate(
            total=Sum(F('valor_original') + F('valor_juros') - F('valor_desconto') - F('valor_pago'))
        )['total'] or 0,
    }
    
    # Dados para filtros
    categorias = CategoriaContaPagar.objects.filter(empresa=empresa, ativo=True).order_by('nome')
    
    context = {
        'contas': contas,
        'stats': stats,
        'categorias': categorias,
        'status_choices': StatusContaPagar.choices,
        'current_filters': {
            'status': status_filter,
            'categoria': categoria_filter,
            'subcategoria': subcategoria_filter,
            'fornecedor': fornecedor_filter,
            'venc_de': vencimento_de,
            'venc_ate': vencimento_ate,
            'search': search,
        },
    }
    
    return render(request, 'financeiro/contas_pagar_listar.html', context)


@login_required
def conta_pagar_nova(request):
    """Criar nova conta a pagar"""
    empresa = get_current_empresa()
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Criar conta
                conta = ContaPagar.objects.create(
                    empresa=empresa,
                    categoria_id=request.POST.get('categoria'),
                    subcategoria_id=request.POST.get('subcategoria'),
                    fornecedor_nome=request.POST.get('fornecedor_nome'),
                    fornecedor_documento=request.POST.get('fornecedor_documento', ''),
                    numero_documento=request.POST.get('numero_documento'),
                    tipo_documento=request.POST.get('tipo_documento'),
                    descricao=request.POST.get('descricao'),
                    valor_original=Decimal(request.POST.get('valor_original')),
                    valor_juros=Decimal(request.POST.get('valor_juros', 0)),
                    valor_desconto=Decimal(request.POST.get('valor_desconto', 0)),
                    data_emissao=request.POST.get('data_emissao'),
                    data_vencimento=request.POST.get('data_vencimento'),
                    parcelado=request.POST.get('parcelado') == 'on',
                    numero_parcelas=int(request.POST.get('numero_parcelas', 1)),
                    recorrente=request.POST.get('recorrente') == 'on',
                    centro_custo=request.POST.get('centro_custo', ''),
                    observacoes=request.POST.get('observacoes', ''),
                    usuario_criacao=request.user
                )
                
                # Se for parcelada, gerar parcelas
                if conta.parcelado and conta.numero_parcelas > 1:
                    conta.gerar_parcelas()
                
                messages.success(request, f'Conta {conta.numero_documento} criada com sucesso!')
                return redirect('financeiro:conta_pagar_detalhes', conta_id=conta.id)
                
        except Exception as e:
            messages.error(request, f'Erro ao criar conta: {str(e)}')
    
    # Dados para o formulário
    categorias = CategoriaContaPagar.objects.filter(empresa=empresa, ativo=True).order_by('nome')
    ultimas_contas = ContaPagar.objects.filter(
        empresa=empresa
    ).select_related('categoria', 'subcategoria').order_by('-created_at')[:5]
    
    context = {
        'categorias': categorias,
        'tipo_documento_choices': TipoDocumento.choices,
        'hoje': timezone.now().date(),
        'ultimas_contas': ultimas_contas,
    }
    
    return render(request, 'financeiro/conta_pagar_nova.html', context)


@login_required
def conta_pagar_detalhes(request, conta_id):
    """Detalhes de uma conta a pagar"""
    conta = get_object_or_404(
        ContaPagar.objects.select_related('categoria', 'subcategoria', 'usuario_criacao'),
        id=conta_id,
        empresa=get_current_empresa()
    )
    
    # Buscar parcelas se houver
    parcelas = conta.parcelas.all().order_by('numero_parcela') if conta.parcelado else None
    
    # Histórico de pagamentos
    pagamentos = conta.pagamentos_efetuados.all().order_by('-data_pagamento')
    
    context = {
        'conta': conta,
        'parcelas': parcelas,
        'pagamentos': pagamentos,
    }
    
    return render(request, 'financeiro/conta_pagar_detalhes.html', context)


@login_required
def registrar_pagamento_conta_pagar(request, conta_id):
    """Registrar pagamento de uma conta a pagar"""
    conta = get_object_or_404(ContaPagar, id=conta_id, empresa=get_current_empresa())
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                valor = Decimal(request.POST.get('valor'))
                data_pagamento = request.POST.get('data_pagamento')
                forma_pagamento = request.POST.get('forma_pagamento')
                parcela_id = request.POST.get('parcela_id')
                
                # Validações
                if valor <= 0:
                    raise ValueError("Valor deve ser maior que zero")
                
                if valor > conta.valor_pendente:
                    raise ValueError("Valor não pode ser maior que o pendente")
                
                # Buscar parcela se especificada
                parcela = None
                if parcela_id:
                    parcela = get_object_or_404(ParcelaContaPagar, id=parcela_id, conta_pagar=conta)
                
                # Criar pagamento
                pagamento = PagamentoEfetuado.objects.create(
                    empresa=get_current_empresa(),
                    conta_pagar=conta,
                    parcela=parcela,
                    valor=valor,
                    data_pagamento=data_pagamento,
                    forma_pagamento=forma_pagamento,
                    banco=request.POST.get('banco', ''),
                    agencia=request.POST.get('agencia', ''),
                    conta=request.POST.get('conta', ''),
                    numero_transacao=request.POST.get('numero_transacao', ''),
                    observacoes=request.POST.get('observacoes', ''),
                    usuario_pagamento=request.user
                )
                
                messages.success(request, f'Pagamento de R$ {valor:.2f} registrado com sucesso!')
                return redirect('financeiro:conta_pagar_detalhes', conta_id=conta.id)
                
        except Exception as e:
            messages.error(request, f'Erro ao registrar pagamento: {str(e)}')
    
    return redirect('financeiro:conta_pagar_detalhes', conta_id=conta.id)


# =============================================================================
# CATEGORIAS E SUBCATEGORIAS
# =============================================================================

@login_required
def categorias_listar(request):
    """Lista e gerencia categorias e subcategorias"""
    empresa = get_current_empresa()
    categorias = CategoriaContaPagar.objects.filter(
        empresa=empresa
    ).prefetch_related('subcategorias').order_by('nome')
    
    context = {
        'categorias': categorias,
    }
    
    return render(request, 'financeiro/categorias_listar.html', context)


@login_required
def categoria_nova(request):
    """Criar nova categoria"""
    if request.method == 'POST':
        try:
            categoria = CategoriaContaPagar.objects.create(
                empresa=get_current_empresa(),
                nome=request.POST.get('nome'),
                descricao=request.POST.get('descricao', ''),
                codigo=request.POST.get('codigo', ''),
                dedutivel=request.POST.get('dedutivel') == 'on',
                centro_custo=request.POST.get('centro_custo', ''),
            )
            
            messages.success(request, f'Categoria {categoria.nome} criada com sucesso!')
            return redirect('financeiro:categorias_listar')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar categoria: {str(e)}')
    
    return render(request, 'financeiro/categoria_nova.html')


@login_required
def subcategoria_nova(request, categoria_id):
    """Criar nova subcategoria"""
    categoria = get_object_or_404(
        CategoriaContaPagar,
        id=categoria_id,
        empresa=get_current_empresa()
    )
    
    if request.method == 'POST':
        try:
            subcategoria = SubcategoriaContaPagar.objects.create(
                empresa=get_current_empresa(),
                categoria=categoria,
                nome=request.POST.get('nome'),
                descricao=request.POST.get('descricao', ''),
                codigo=request.POST.get('codigo', ''),
                recorrente=request.POST.get('recorrente') == 'on',
                dia_vencimento_padrao=int(request.POST.get('dia_vencimento_padrao')) if request.POST.get('dia_vencimento_padrao') else None,
            )
            
            messages.success(request, f'Subcategoria {subcategoria.nome} criada com sucesso!')
            return redirect('financeiro:categorias_listar')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar subcategoria: {str(e)}')
    
    context = {
        'categoria': categoria,
    }
    
    return render(request, 'financeiro/subcategoria_nova.html', context)


# =============================================================================
# AJAX VIEWS - CONTAS A PAGAR
# =============================================================================

@login_required
def buscar_subcategorias_ajax(request):
    """Buscar subcategorias de uma categoria via AJAX"""
    categoria_id = request.GET.get('categoria_id')
    subcategorias = []
    
    if categoria_id:
        subcategorias_obj = SubcategoriaContaPagar.objects.filter(
            empresa=get_current_empresa(),
            categoria_id=categoria_id,
            ativo=True
        ).order_by('nome')
        
        subcategorias = [
            {
                'id': s.id,
                'nome': s.nome,
                'recorrente': s.recorrente,
                'dia_vencimento': s.dia_vencimento_padrao
            }
            for s in subcategorias_obj
        ]
    
    return JsonResponse({'subcategorias': subcategorias})


@login_required
def buscar_fornecedores_ajax(request):
    """Buscar fornecedores via AJAX para autocomplete"""
    query = request.GET.get('q', '')
    fornecedores = []
    
    if query:
        # Buscar fornecedores únicos das contas existentes
        fornecedores_obj = ContaPagar.objects.filter(
            empresa=get_current_empresa(),
            fornecedor_nome__icontains=query
        ).values('fornecedor_nome', 'fornecedor_documento').distinct()[:10]
        
        fornecedores = [
            {
                'nome': f['fornecedor_nome'],
                'documento': f['fornecedor_documento'] or ''
            }
            for f in fornecedores_obj
        ]
    
    return JsonResponse({'fornecedores': fornecedores})


# =============================================================================
# RELATÓRIOS FINANCEIROS INTEGRADOS
# =============================================================================

@login_required
def relatorio_dre(request):
    """Demonstrativo de Resultados do Exercício (DRE)"""
    empresa = get_current_empresa()
    
    # Período do relatório
    mes = int(request.GET.get('mes', timezone.now().month))
    ano = int(request.GET.get('ano', timezone.now().year))
    
    # Buscar dados de receitas (contas a receber)
    inicio_periodo = date(ano, mes, 1)
    if mes == 12:
        fim_periodo = date(ano + 1, 1, 1) - timedelta(days=1)
    else:
        fim_periodo = date(ano, mes + 1, 1) - timedelta(days=1)
    
    # Receitas
    receitas = PagamentoRecebido.objects.filter(
        empresa=empresa,
        data_pagamento__gte=inicio_periodo,
        data_pagamento__lte=fim_periodo
    ).aggregate(total=Sum('valor_pago'))['total'] or 0
    
    # Despesas por categoria
    despesas_por_categoria = {}
    categorias = CategoriaContaPagar.objects.filter(empresa=empresa, ativo=True)
    
    for categoria in categorias:
        valor = PagamentoEfetuado.objects.filter(
            empresa=empresa,
            conta_pagar__categoria=categoria,
            data_pagamento__gte=inicio_periodo,
            data_pagamento__lte=fim_periodo
        ).aggregate(total=Sum('valor'))['total'] or 0
        
        if valor > 0:
            despesas_por_categoria[categoria.nome] = valor
    
    total_despesas = sum(despesas_por_categoria.values())
    resultado = receitas - total_despesas
    
    context = {
        'mes': mes,
        'ano': ano,
        'receitas': receitas,
        'despesas_por_categoria': despesas_por_categoria,
        'total_despesas': total_despesas,
        'resultado': resultado,
        'meses': [
            (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'),
            (4, 'Abril'), (5, 'Maio'), (6, 'Junho'),
            (7, 'Julho'), (8, 'Agosto'), (9, 'Setembro'),
            (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro')
        ],
        'anos_disponiveis': range(2020, timezone.now().year + 2),
    }
    
    return render(request, 'financeiro/relatorio_dre.html', context)


@login_required
def relatorio_fluxo_caixa(request):
    """Relatório de Fluxo de Caixa"""
    empresa = get_current_empresa()
    hoje = timezone.now().date()
    
    # Período do relatório (próximos 30 dias por padrão)
    dias = int(request.GET.get('dias', 30))
    data_fim = hoje + timedelta(days=dias)
    
    # Buscar contas a receber no período
    contas_receber = ContaReceber.objects.filter(
        empresa=empresa,
        data_vencimento__gte=hoje,
        data_vencimento__lte=data_fim,
        status__in=[StatusPagamento.PENDENTE, StatusPagamento.PARCIAL]
    ).order_by('data_vencimento')
    
    # Buscar contas a pagar no período
    contas_pagar = ContaPagar.objects.filter(
        empresa=empresa,
        data_vencimento__gte=hoje,
        data_vencimento__lte=data_fim,
        status__in=[StatusContaPagar.PENDENTE, StatusContaPagar.PARCIAL]
    ).order_by('data_vencimento')
    
    # Agrupar por dia
    fluxo_por_dia = {}
    
    # Adicionar recebimentos
    for conta in contas_receber:
        data = conta.data_vencimento
        if data not in fluxo_por_dia:
            fluxo_por_dia[data] = {'entradas': 0, 'saidas': 0, 'saldo': 0}
        fluxo_por_dia[data]['entradas'] += conta.valor_pendente
    
    # Adicionar pagamentos
    for conta in contas_pagar:
        data = conta.data_vencimento
        if data not in fluxo_por_dia:
            fluxo_por_dia[data] = {'entradas': 0, 'saidas': 0, 'saldo': 0}
        fluxo_por_dia[data]['saidas'] += conta.valor_pendente
    
    # Calcular saldo acumulado
    saldo_acumulado = 0
    fluxo_ordenado = []
    
    for data in sorted(fluxo_por_dia.keys()):
        dia = fluxo_por_dia[data]
        dia['saldo'] = dia['entradas'] - dia['saidas']
        saldo_acumulado += dia['saldo']
        dia['saldo_acumulado'] = saldo_acumulado
        dia['data'] = data
        fluxo_ordenado.append(dia)
    
    context = {
        'fluxo_caixa': fluxo_ordenado,
        'total_entradas': sum(d['entradas'] for d in fluxo_ordenado),
        'total_saidas': sum(d['saidas'] for d in fluxo_ordenado),
        'saldo_final': saldo_acumulado,
        'dias': dias,
        'data_inicio': hoje,
        'data_fim': data_fim,
    }
    
    return render(request, 'financeiro/relatorio_fluxo_caixa.html', context)
