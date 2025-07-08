from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum, Count, Avg, F
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.core.paginator import Paginator
from django.db import transaction
from django.utils import timezone
from datetime import date, datetime, timedelta
import json

# DRF imports
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend

from .models import (
    OrdemProducao, GradeProducao, Departamento, MateriaPrima,
    ConsumoMateriaPrima, ProcessoProducao, CapacidadeProducao,
    RelatorioFaturamento, StatusOP, LinhaProducao, EtapaProducao, ControleEtapaOP,
    HistoricoProducao, StatusLinha
)
from .serializers import (
    OrdemProducaoListSerializer, OrdemProducaoDetailSerializer,
    OrdemProducaoCreateUpdateSerializer, DepartamentoSerializer,
    MateriaPrimaSerializer, CapacidadeProducaoSerializer,
    RelatorioFaturamentoSerializer, DashboardStatsSerializer,
    BulkUpdateStatusSerializer, BulkUpdateProgressSerializer,
    ProcessoProducaoSerializer, GradeProducaoSerializer
)
from apps.core.mixins import TenantMixin
from apps.core.middleware import get_current_empresa


@login_required
def faturamento_view(request):
    """
    View principal de Faturamento - Replica exatamente a planilha
    Interface idêntica à primeira imagem mostrada
    """
    empresa = get_current_empresa()
    
    # Filtros da interface
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    cliente_filtro = request.GET.get('cliente')
    status_filtro = request.GET.get('status')
    
    # Query base das OPs
    ops = OrdemProducao.objects.filter(empresa=empresa)
    
    # Aplicar filtros
    if data_inicio:
        ops = ops.filter(data_entrada__gte=data_inicio)
    if data_fim:
        ops = ops.filter(data_entrada__lte=data_fim)
    if cliente_filtro:
        ops = ops.filter(cliente__nome__icontains=cliente_filtro)
    if status_filtro:
        ops = ops.filter(status=status_filtro)
    
    ops = ops.order_by('-data_entrada', 'numero_op')
    
    # Dados para o cabeçalho (como na planilha)
    try:
        capacidade = CapacidadeProducao.objects.get(empresa=empresa)
    except CapacidadeProducao.DoesNotExist:
        capacidade = CapacidadeProducao.objects.create(
            empresa=empresa,
            capacidade_diaria=300
        )
    
    # Totalizadores (como na planilha)
    totais = ops.aggregate(
        total_quantidade=Sum('quantidade'),
        total_valor=Sum('preco_total'),
        total_ops=Count('id')
    )
    
    # Estatísticas por status
    stats_status = {}
    for status_code, status_name in StatusOP.choices:
        count = ops.filter(status=status_code).count()
        valor = ops.filter(status=status_code).aggregate(
            total=Sum('preco_total')
        )['total'] or 0
        stats_status[status_code] = {
            'count': count,
            'valor': valor,
            'nome': status_name
        }
    
    context = {
        'ops': ops,
        'capacidade': capacidade,
        'data_atual': date.today(),
        'totais': totais,
        'stats_status': stats_status,
        'filtros': {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'cliente': cliente_filtro,
            'status': status_filtro,
        },
        'status_choices': StatusOP.choices,
    }
    
    return render(request, 'producao/faturamento.html', context)


@login_required
def relatorio_faturamento_view(request):
    """
    View do Relatório de Faturamento - Replica a segunda imagem
    Mostra dados consolidados por empresa e período
    """
    empresa = get_current_empresa()
    
    # Parâmetros do relatório
    mes = int(request.GET.get('mes', timezone.now().month))
    ano = int(request.GET.get('ano', timezone.now().year))
    
    # Período do relatório
    data_inicio = date(ano, mes, 1)
    if mes == 12:
        data_fim = date(ano + 1, 1, 1)
    else:
        data_fim = date(ano, mes + 1, 1)
    
    # OPs do período
    ops_periodo = OrdemProducao.objects.filter(
        empresa=empresa,
        data_entrada__gte=data_inicio,
        data_entrada__lt=data_fim
    )
    
    # Dados consolidados (como na planilha)
    entradas = ops_periodo.filter(status='CADASTRADA').count()
    saidas = ops_periodo.filter(status='CONCLUIDA').count()
    a_produzir = ops_periodo.filter(status__in=['CADASTRADA', 'EM_PRODUCAO']).count()
    
    # Valores financeiros
    valor_entradas = ops_periodo.aggregate(
        total=Sum('preco_total')
    )['total'] or 0
    
    valor_saidas = ops_periodo.filter(status='CONCLUIDA').aggregate(
        total=Sum('preco_total')
    )['total'] or 0
    
    # Simular valores recebidos (70% do faturado)
    valor_recebido = valor_saidas * 0.7
    falta_receber = valor_saidas - valor_recebido
    
    # Dados por cliente (como na planilha)
    clientes_dados = []
    for op in ops_periodo.select_related('cliente'):
        cliente_info = {
            'op': op.numero_op,
            'cliente': op.cliente.nome,
            'quantidade': op.quantidade_total,
            'preco_unitario': op.preco_unitario,
            'preco_total': op.preco_total,
            'data_pagamento': op.data_conclusao,
            'valor_pago': op.preco_total if op.status == 'ENTREGUE' else 0,
            'referencia': f"{op.data_entrada.strftime('%m-%Y')} - {empresa.nome}",
        }
        clientes_dados.append(cliente_info)
    
    # Mensagem de alerta (como na planilha)
    # "Atenção, vamos trabalhar porque o negócio esta feio!"
    percentual_ocupacao = (entradas / (capacidade.capacidade_diaria * 22)) * 100 if hasattr(locals(), 'capacidade') else 50
    
    if percentual_ocupacao < 70:
        alerta = "Atenção, vamos trabalhar porque o negócio está feio!"
        alerta_tipo = "danger"
    elif percentual_ocupacao < 85:
        alerta = "Produção estável, mas podemos melhorar!"
        alerta_tipo = "warning"
    else:
        alerta = "Excelente! Produção em alta!"
        alerta_tipo = "success"
    
    context = {
        'mes': mes,
        'ano': ano,
        'empresa': empresa,
        'entradas': entradas,
        'saidas': saidas,
        'a_produzir': a_produzir,
        'valor_entradas': valor_entradas,
        'valor_saidas': valor_saidas,
        'valor_recebido': valor_recebido,
        'falta_receber': falta_receber,
        'clientes_dados': clientes_dados,
        'alerta': alerta,
        'alerta_tipo': alerta_tipo,
        'percentual_ocupacao': percentual_ocupacao,
    }
    
    return render(request, 'producao/relatorio_faturamento.html', context)


@login_required
def ajax_op_detalhes(request, op_id):
    """Ajax para carregar detalhes de uma OP"""
    op = get_object_or_404(OrdemProducao, id=op_id, empresa=get_current_empresa())
    
    data = {
        'numero_op': op.numero_op,
        'op_externa': op.op_externa,
        'produto': f"{op.produto.codigo} - {op.produto.referencia}" if op.produto else '',
        'cliente': op.cliente.nome if op.cliente else '',
        'quantidade': op.quantidade_total,
        'preco_unitario': str(op.preco_unitario) if op.preco_unitario else '0',
        'preco_total': str(op.preco_total) if hasattr(op, 'preco_total') else '0',
        'data_entrada': op.data_entrada.strftime('%d/%m/%Y') if op.data_entrada else '',
        'data_previsao': op.data_previsao.strftime('%d/%m/%Y') if op.data_previsao else '',
        'status': op.get_status_display(),
        'status_color': getattr(op, 'status_color', ''),
        'observacoes': op.observacoes,
    }
    
    return JsonResponse(data)


@login_required
def ajax_capacidade_producao(request):
    """Ajax para atualizar capacidade de produção"""
    if request.method == 'POST':
        nova_capacidade = int(request.POST.get('capacidade', 300))
        empresa = get_current_empresa()
        
        capacidade, created = CapacidadeProducao.objects.get_or_create(
            empresa=empresa,
            defaults={'capacidade_diaria': nova_capacidade}
        )
        
        if not created:
            capacidade.capacidade_diaria = nova_capacidade
            capacidade.save()
        
        return JsonResponse({
            'success': True,
            'capacidade_diaria': capacidade.capacidade_diaria,
            'capacidade_mensal': capacidade.capacidade_mensal
        })
    
    return JsonResponse({'success': False})


# =============================================================================
# VIEWS DO MÓDULO DE ORDENS DE PRODUÇÃO
# =============================================================================

@login_required
def dashboard_producao(request):
    """Dashboard principal do módulo de produção"""
    empresa = get_current_empresa()
    
    # Estatísticas básicas
    stats = {
        'ops_ativas': OrdemProducao.objects.filter(empresa=empresa, status__in=['CADASTRADA', 'EM_PRODUCAO']).count(),
        'em_producao': OrdemProducao.objects.filter(empresa=empresa, status='EM_PRODUCAO').count(),
        'atrasadas': OrdemProducao.objects.filter(empresa=empresa, data_previsao__lt=timezone.now().date(), status__in=['CADASTRADA', 'EM_PRODUCAO']).count(),
        'pecas_dia': 1850,  # Valor temporário
    }
    
    context = {
        'stats': stats,
        'empresa': empresa,
    }
    
    return render(request, 'producao/dashboard_producao.html', context)


@login_required
def ops_listar(request):
    """Lista OPs com filtros e busca - Dados reais com workflow"""
    empresa = get_current_empresa()
    ops = OrdemProducao.objects.filter(empresa=empresa)
    
    # Filtros
    status_filter = request.GET.get('status')
    prioridade_filter = request.GET.get('prioridade')
    data_entrega_filter = request.GET.get('data_entrega')
    search = request.GET.get('search')
    
    if status_filter:
        ops = ops.filter(status=status_filter)
    if prioridade_filter:
        ops = ops.filter(prioridade=prioridade_filter)
    if data_entrega_filter:
        ops = ops.filter(data_previsao__lte=data_entrega_filter)
    if search:
        ops = ops.filter(
            Q(numero_op__icontains=search) |
            Q(op_externa__icontains=search) |
            Q(produto__referencia__icontains=search) |
            Q(produto__codigo__icontains=search) |
            Q(cliente__nome__icontains=search)
        )
    
    # Ordenação
    ops = ops.select_related('cliente', 'produto', 'responsavel').order_by('-data_entrada')
    
    # Adicionar informações do workflow para cada OP
    from .workflow import WorkflowOP
    ops_com_workflow = []
    for op in ops:
        workflow = WorkflowOP(op)
        op.proximos_status = workflow.obter_proximos_status(op.status)
        op.pode_avancar = len(op.proximos_status) > 0 and op.proximos_status[0] != 'CANCELADA'
        op.cor_status = workflow.obter_cor_status(op.status)
        ops_com_workflow.append(op)
    
    # Paginação
    paginator = Paginator(ops_com_workflow, 20)
    page = request.GET.get('page')
    ops_page = paginator.get_page(page)
    
    # Estatísticas completas para cards - usando dados reais
    all_ops = OrdemProducao.objects.filter(empresa=empresa)
    stats = {
        'total': all_ops.count(),
        'planejamento': all_ops.filter(status='CADASTRADA').count(),
        'preparacao': all_ops.filter(status='PREPARACAO').count(),
        'frente_externa': all_ops.filter(status='FRENTE_EXTERNA').count(),
        'montagem': all_ops.filter(status='MONTAGEM').count(),
        'em_producao': all_ops.filter(status='EM_PRODUCAO').count(),
        'concluidas': all_ops.filter(status='CONCLUIDA').count(),
        'finalizadas': all_ops.filter(status='FINALIZADA').count(),
        'entregues': all_ops.filter(status='ENTREGUE').count(),
        'canceladas': all_ops.filter(status='CANCELADA').count(),
        'atrasadas': all_ops.filter(
            data_previsao__lt=date.today(),
            status__in=['CADASTRADA', 'PREPARACAO', 'FRENTE_EXTERNA', 'MONTAGEM', 'EM_PRODUCAO']
        ).count(),
    }
    
    # Status choices expandido
    status_choices = [
        ('CADASTRADA', 'OP Cadastrada'),
        ('PREPARACAO', 'Preparação'),
        ('FRENTE_EXTERNA', 'Frente Externa'),
        ('MONTAGEM', 'Montagem'),
        ('EM_PRODUCAO', 'Em Produção'),
        ('CONCLUIDA', 'Concluída'),
        ('FINALIZADA', 'Finalizada'),
        ('ENTREGUE', 'Entregue'),
        ('CANCELADA', 'Cancelada'),
    ]
    
    context = {
        'ops': ops_page,
        'stats': stats,
        'empresa': empresa,
        'status_choices': status_choices,
        'current_filters': {
            'status': status_filter,
            'prioridade': prioridade_filter,
            'data_entrega': data_entrega_filter,
            'search': search,
        },
        'is_paginated': paginator.num_pages > 1,
        'page_obj': ops_page,
    }
    return render(request, 'producao/ops_listar.html', context)


@login_required
def op_form(request, op_id=None):
    """Formulário de criação/edição de OP"""
    op = None
    
    if op_id:
        op = get_object_or_404(OrdemProducao, id=op_id, empresa=request.empresa_atual)
    
    if request.method == 'POST':
        # Processar formulário
        try:
            with transaction.atomic():
                # Dados básicos da OP
                data_previsao = request.POST.get('data_previsao')
                
                op_data = {
                    'op_externa': request.POST.get('op_externa'),
                    'data_previsao': data_previsao,
                    'preco_unitario': request.POST.get('preco_unitario'),
                    'prioridade': request.POST.get('prioridade', 1),
                    'observacoes': request.POST.get('observacoes', ''),
                    'cliente_id': request.POST.get('cliente'),
                    'produto_id': request.POST.get('produto')
                }
                
                if op:
                    # Atualizar OP existente
                    for key, value in op_data.items():
                        setattr(op, key, value)
                    op.save()
                else:
                    # Criar nova OP
                    op_data['empresa'] = request.empresa_atual
                    op = OrdemProducao.objects.create(**op_data)
                
                # Processar grade de tamanhos
                op.itens_grade.all().delete()
                
                tamanhos = request.POST.getlist('tamanhos[]')
                quantidades = request.POST.getlist('quantidades[]')

                for tamanho, quantidade in zip(tamanhos, quantidades):
                    if tamanho and quantidade and int(quantidade) > 0:
                        GradeProducao.objects.create(
                            ordem_producao=op,
                            tamanho=tamanho,
                            quantidade=int(quantidade)
                        )
                
                messages.success(request, f'OP {op.numero_op} salva com sucesso!')
                return redirect('producao:op_detalhes', op_id=op.id)
        
        except Exception as e:
            messages.error(request, f'Erro ao salvar OP: {str(e)}')
    
    context = {
        'op': op,
        'empresa': request.empresa_atual,
        'is_edit': op is not None,
    }
    return render(request, 'producao/op_form.html', context)


@login_required
def op_detalhes(request, op_id):
    """Detalhes da OP com timeline, workflow e dados reais"""
    op = get_object_or_404(
        OrdemProducao.objects.select_related('cliente', 'produto', 'responsavel'), 
        id=op_id, 
        empresa=get_current_empresa()
    )
    
    # Informações do workflow
    from .workflow import WorkflowOP
    workflow = WorkflowOP(op)
    op.proximos_status = workflow.obter_proximos_status(op.status)
    op.pode_avancar = len(op.proximos_status) > 0 and op.proximos_status[0] != 'CANCELADA'
    op.cor_status = workflow.obter_cor_status(op.status)
    
    # Timeline de status (simulado - pode ser implementado com histórico real)
    timeline_status = [
        {'status': 'CADASTRADA', 'nome': 'OP Cadastrada', 'concluido': True, 'ativo': op.status == 'CADASTRADA'},
        {'status': 'PREPARACAO', 'nome': 'Preparação', 'concluido': op.status in ['PREPARACAO', 'FRENTE_EXTERNA', 'MONTAGEM', 'EM_PRODUCAO', 'CONCLUIDA', 'FINALIZADA', 'ENTREGUE'], 'ativo': op.status == 'PREPARACAO'},
        {'status': 'FRENTE_EXTERNA', 'nome': 'Frente Externa', 'concluido': op.status in ['FRENTE_EXTERNA', 'MONTAGEM', 'EM_PRODUCAO', 'CONCLUIDA', 'FINALIZADA', 'ENTREGUE'], 'ativo': op.status == 'FRENTE_EXTERNA'},
        {'status': 'MONTAGEM', 'nome': 'Montagem', 'concluido': op.status in ['MONTAGEM', 'EM_PRODUCAO', 'CONCLUIDA', 'FINALIZADA', 'ENTREGUE'], 'ativo': op.status == 'MONTAGEM'},
        {'status': 'EM_PRODUCAO', 'nome': 'Em Produção', 'concluido': op.status in ['EM_PRODUCAO', 'CONCLUIDA', 'FINALIZADA', 'ENTREGUE'], 'ativo': op.status == 'EM_PRODUCAO'},
        {'status': 'CONCLUIDA', 'nome': 'Concluída', 'concluido': op.status in ['CONCLUIDA', 'FINALIZADA', 'ENTREGUE'], 'ativo': op.status == 'CONCLUIDA'},
        {'status': 'FINALIZADA', 'nome': 'Finalizada', 'concluido': op.status in ['FINALIZADA', 'ENTREGUE'], 'ativo': op.status == 'FINALIZADA'},
        {'status': 'ENTREGUE', 'nome': 'Entregue', 'concluido': op.status == 'ENTREGUE', 'ativo': op.status == 'ENTREGUE'},
    ]
    
    # Calcular progresso geral
    status_order = ['CADASTRADA', 'PREPARACAO', 'FRENTE_EXTERNA', 'MONTAGEM', 'EM_PRODUCAO', 'CONCLUIDA', 'FINALIZADA', 'ENTREGUE']
    if op.status in status_order:
        progresso_geral = ((status_order.index(op.status) + 1) / len(status_order)) * 100
    else:
        progresso_geral = 0
    
    # Grade de produção com dados reais
    grade = op.itens_grade.all().order_by('tamanho')
    
    # Matérias primas por departamento (dados reais se existirem)
    materias_por_depto = {}
    for consumo in op.materias_primas.select_related('materia_prima', 'departamento'):
        depto = consumo.departamento.nome
        if depto not in materias_por_depto:
            materias_por_depto[depto] = []
        materias_por_depto[depto].append(consumo)
    
    # Processos de produção (dados reais se existirem)
    processos = op.processos.select_related('departamento', 'responsavel').order_by('departamento__ordem')
    
    # Etapas de controle (se existirem)
    etapas_controle = []
    if hasattr(op, 'controles_etapa'):
        etapas_controle = op.controles_etapa.select_related('etapa', 'linha_producao', 'responsavel').order_by('etapa__ordem')
    
    context = {
        'op': op,
        'empresa': get_current_empresa(),
        'grade': grade,
        'materias_por_depto': materias_por_depto,
        'processos': processos,
        'etapas_controle': etapas_controle,
        'timeline_status': timeline_status,
        'progresso_geral': progresso_geral,
        'workflow': workflow,
    }
    return render(request, 'producao/op_detalhes.html', context)


@login_required
def op_pdf(request, op_id):
    """Gerar PDF da OP"""
    op = get_object_or_404(OrdemProducao, id=op_id, empresa=request.empresa_atual)
    
    # TODO: Implementar geração de PDF com ReportLab
    # Por enquanto, retornar resposta simples
    return HttpResponse(f"PDF da OP #{op.numero_op} será gerado aqui", content_type='text/plain')


@login_required
def linhas_producao(request):
    """Gestão de linhas de produção"""
    empresa = request.empresa_atual
    
    # Buscar linhas de produção
    linhas = LinhaProducao.objects.filter(empresa=empresa, ativo=True).prefetch_related(
        'operadores', 'ops_linha'
    )
    
    # Estatísticas gerais
    total_linhas = linhas.count()
    linhas_ativas = linhas.filter(status='ATIVA').count()
    linhas_paradas = linhas.filter(status='PARADA').count()
    linhas_manutencao = linhas.filter(status='MANUTENCAO').count()
    
    # Produção total hoje
    hoje = timezone.now().date()
    producao_hoje = sum(linha.producao_hoje for linha in linhas)
    capacidade_total = sum(linha.capacidade_diaria for linha in linhas)
    eficiencia_geral = (producao_hoje / capacidade_total * 100) if capacidade_total > 0 else 0
    
    # OPs em produção por linha
    ops_por_linha = {}
    for linha in linhas:
        ops_por_linha[linha.id] = {
            'ops_ativas': linha.ops_ativas,
            'producao_hoje': linha.producao_hoje,
            'eficiencia_hoje': linha.eficiencia_hoje,
            'capacidade_diaria': linha.capacidade_diaria
        }
    
    context = {
        'empresa': empresa,
        'linhas': linhas,
        'ops_por_linha': ops_por_linha,
        'stats': {
            'total_linhas': total_linhas,
            'linhas_ativas': linhas_ativas,
            'linhas_paradas': linhas_paradas,
            'linhas_manutencao': linhas_manutencao,
            'producao_hoje': producao_hoje,
            'capacidade_total': capacidade_total,
            'eficiencia_geral': round(eficiencia_geral, 1)
        }
    }
    return render(request, 'producao/linhas_producao.html', context)


@login_required
def linha_detalhes(request, linha_id):
    """Detalhes de uma linha de produção com controle de etapas"""
    empresa = request.empresa_atual
    linha = get_object_or_404(LinhaProducao, id=linha_id, empresa=empresa)
    
    # OPs atualmente na linha
    ops_ativas = linha.ops_linha.filter(
        status__in=['CADASTRADA', 'EM_PRODUCAO']
    ).select_related('cliente', 'produto').prefetch_related('controles_etapa__etapa')
    
    # Etapas configuradas para a empresa
    etapas = EtapaProducao.objects.filter(empresa=empresa, ativo=True).order_by('ordem')
    
    # Histórico de produção recente (últimos 7 dias)
    sete_dias_atras = timezone.now() - timezone.timedelta(days=7)
    historico_recente = HistoricoProducao.objects.filter(
        linha_producao=linha,
        data_registro__gte=sete_dias_atras
    ).select_related('ordem_producao', 'etapa', 'operador').order_by('-data_registro')[:20]
    
    # Estatísticas da linha (últimos 30 dias)
    trinta_dias_atras = timezone.now() - timezone.timedelta(days=30)
    stats_30_dias = HistoricoProducao.objects.filter(
        linha_producao=linha,
        data_registro__gte=trinta_dias_atras
    ).aggregate(
        total_produzido=Sum('quantidade_produzida'),
        total_defeituoso=Sum('quantidade_defeituosa'),
        total_retrabalho=Sum('quantidade_retrabalho'),
        tempo_total=Sum('tempo_producao_minutos')
    )
    
    # Eficiência por etapa
    eficiencia_etapas = {}
    for etapa in etapas:
        historicos_etapa = HistoricoProducao.objects.filter(
            linha_producao=linha,
            etapa=etapa,
            data_registro__gte=trinta_dias_atras
        )
        
        if historicos_etapa.exists():
            tempo_medio = historicos_etapa.aggregate(
                tempo_medio=Avg('tempo_producao_minutos')
            )['tempo_medio'] or 0
            
            eficiencia = (etapa.tempo_medio_minutos / tempo_medio * 100) if tempo_medio > 0 else 0
            eficiencia_etapas[etapa.id] = min(100, max(0, eficiencia))
        else:
            eficiencia_etapas[etapa.id] = 0
    
    # Gráfico de produção diária (últimos 15 dias)
    producao_diaria = []
    for i in range(15):
        data = timezone.now().date() - timezone.timedelta(days=i)
        producao_dia = HistoricoProducao.objects.filter(
            linha_producao=linha,
            data_registro__date=data
        ).aggregate(total=Sum('quantidade_produzida'))['total'] or 0
        
        producao_diaria.append({
            'data': data.strftime('%d/%m'),
            'producao': producao_dia
        })
    
    producao_diaria.reverse()  # Ordem cronológica
    
    context = {
        'empresa': empresa,
        'linha': linha,
        'ops_ativas': ops_ativas,
        'etapas': etapas,
        'historico_recente': historico_recente,
        'stats_30_dias': stats_30_dias,
        'eficiencia_etapas': eficiencia_etapas,
        'producao_diaria': producao_diaria
    }
    return render(request, 'producao/linha_detalhes.html', context)


@login_required
def relatorios_producao(request):
    """Página de relatórios de produção com dados dinâmicos"""
    empresa = get_current_empresa()
    hoje = timezone.now().date()
    
    # Filtros de período
    periodo = request.GET.get('periodo', 'mes')
    
    # Calcular datas baseadas no período
    if periodo == 'hoje':
        data_inicio = hoje
        data_fim = hoje
    elif periodo == 'semana':
        data_inicio = hoje - timedelta(days=7)
        data_fim = hoje
    elif periodo == 'mes':
        data_inicio = hoje.replace(day=1)
        data_fim = hoje
    elif periodo == 'trimestre':
        data_inicio = hoje - timedelta(days=90)
        data_fim = hoje
    elif periodo == 'ano':
        data_inicio = hoje.replace(month=1, day=1)
        data_fim = hoje
    else:
        data_inicio = hoje - timedelta(days=30)
        data_fim = hoje
    
    # ========== ESTATÍSTICAS DE PRODUÇÃO ==========
    ops_total = OrdemProducao.objects.filter(
        empresa=empresa,
        data_entrada__gte=data_inicio,
        data_entrada__lte=data_fim
    ).count()
    
    ops_em_producao = OrdemProducao.objects.filter(
        empresa=empresa,
        status=StatusOP.EM_PRODUCAO
    ).count()
    
    ops_concluidas = OrdemProducao.objects.filter(
        empresa=empresa,
        status=StatusOP.CONCLUIDA,
        data_conclusao__gte=data_inicio,
        data_conclusao__lte=data_fim
    ).count()
    
    ops_atrasadas = OrdemProducao.objects.filter(
        empresa=empresa,
        data_previsao__lt=hoje,
        status__in=[StatusOP.CADASTRADA, StatusOP.EM_PRODUCAO]
    ).count()
    
    # ========== EFICIÊNCIA ==========
    # Calcular eficiência baseada em OPs concluídas no prazo
    ops_no_prazo = OrdemProducao.objects.filter(
        empresa=empresa,
        status=StatusOP.CONCLUIDA,
        data_conclusao__gte=data_inicio,
        data_conclusao__lte=data_fim
    ).filter(
        data_conclusao__lte=F('data_previsao')
    ).count()
    
    eficiencia_geral = round((ops_no_prazo / ops_concluidas * 100) if ops_concluidas > 0 else 0, 1)
    
    # Capacidade de utilização
    try:
        capacidade = CapacidadeProducao.objects.get(empresa=empresa)
        capacidade_diaria = capacidade.capacidade_diaria
        dias_uteis = 22  # Média mensal
        capacidade_mensal = capacidade_diaria * dias_uteis
        capacidade_util = round((ops_total / capacidade_mensal * 100) if capacidade_mensal > 0 else 0, 1)
    except:
        capacidade_util = 0
    
    # Qualidade (simulado por enquanto)
    qualidade = 96  # Implementar cálculo real baseado em defeitos/retrabalhos
    
    # ========== FINANCEIRO ==========
    from apps.financeiro.models import ContaReceber, PagamentoRecebido
    
    # Faturamento do período
    faturamento = PagamentoRecebido.objects.filter(
        empresa=empresa,
        data_pagamento__gte=data_inicio,
        data_pagamento__lte=data_fim
    ).aggregate(total=Sum('valor_pago'))['total'] or 0
    
    # Custos (simulado - implementar cálculo real)
    custos = faturamento * 0.7  # 70% do faturamento
    margem = round(((faturamento - custos) / faturamento * 100) if faturamento > 0 else 0, 1)
    
    # ========== MATERIAIS ==========
    from apps.producao.models import MateriaPrima
    
    total_materias = MateriaPrima.objects.filter(
        empresa=empresa,
        ativo=True
    ).count()
    
    baixo_estoque = MateriaPrima.objects.filter(
        empresa=empresa,
        ativo=True,
        estoque_atual__lte=F('estoque_minimo')
    ).count()
    
    # Consumo mensal (simulado)
    consumo_mes = faturamento * 0.3  # 30% do faturamento
    
    # ========== CLIENTES ==========
    from apps.cadastros.models import Cliente
    
    total_clientes = Cliente.objects.filter(empresa=empresa).count()
    
    # Clientes ativos (com OPs no período)
    clientes_ativos = OrdemProducao.objects.filter(
        empresa=empresa,
        data_entrada__gte=data_inicio,
        data_entrada__lte=data_fim
    ).values('cliente').distinct().count()
    
    # Ticket médio
    ticket_medio = faturamento / clientes_ativos if clientes_ativos > 0 else 0
    
    # ========== QUALIDADE ==========
    # Por enquanto valores simulados
    qualidade_index = 94
    defeitos = 3
    retrabalho = 2
    
    # ========== DADOS POR STATUS ==========
    ops_por_status = {}
    for status_code, status_name in StatusOP.choices:
        count = OrdemProducao.objects.filter(
            empresa=empresa,
            status=status_code
        ).count()
        ops_por_status[status_name] = count
    
    # ========== PRODUÇÃO MENSAL (últimos 6 meses) ==========
    producao_mensal = []
    for i in range(6):
        mes_ref = hoje - timedelta(days=i*30)
        inicio_mes = mes_ref.replace(day=1)
        fim_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        ops_mes = OrdemProducao.objects.filter(
            empresa=empresa,
            data_entrada__gte=inicio_mes,
            data_entrada__lte=fim_mes
        ).count()
        
        producao_mensal.append({
            'mes': inicio_mes.strftime('%b/%Y'),
            'quantidade': ops_mes
        })
    
    producao_mensal.reverse()
    
    context = {
        'empresa': empresa,
        'periodo': periodo,
        # Produção
        'ops_total': ops_total,
        'ops_em_producao': ops_em_producao,
        'ops_concluidas': ops_concluidas,
        'ops_atrasadas': ops_atrasadas,
        # Eficiência
        'eficiencia_geral': eficiencia_geral,
        'capacidade_util': capacidade_util,
        'qualidade': qualidade,
        # Financeiro
        'faturamento': faturamento,
        'custos': custos,
        'margem': margem,
        # Materiais
        'total_materias': total_materias,
        'baixo_estoque': baixo_estoque,
        'consumo_mes': consumo_mes,
        # Clientes
        'total_clientes': total_clientes,
        'clientes_ativos': clientes_ativos,
        'ticket_medio': ticket_medio,
        # Qualidade
        'qualidade_index': qualidade_index,
        'defeitos': defeitos,
        'retrabalho': retrabalho,
        # Gráficos
        'ops_por_status': ops_por_status,
        'producao_mensal': producao_mensal,
    }
    
    return render(request, 'producao/relatorios_producao.html', context)


@login_required
def relatorio_producao(request):
    """Relatório de produção"""
    context = {
        'empresa': request.empresa_atual,
    }
    return render(request, 'producao/relatorio_producao.html', context)


@login_required
def relatorio_eficiencia(request):
    """Relatório de eficiência"""
    context = {
        'empresa': request.empresa_atual,
    }
    return render(request, 'producao/relatorio_eficiencia.html', context)


@login_required
def materias_op(request, op_id):
    """Matérias primas de uma OP específica"""
    op = get_object_or_404(OrdemProducao, id=op_id, empresa=request.empresa_atual)
    
    # Matérias primas consumidas por esta OP
    materias = op.materias_primas.select_related('materia_prima', 'departamento').all()
    
    context = {
        'op': op,
        'empresa': request.empresa_atual,
        'materias': materias,
    }
    return render(request, 'producao/materias_op.html', context)


@login_required
def exportar_ops(request):
    """Exportar lista de OPs para Excel"""
    try:
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ordens de Produção"
        
        # Cabeçalhos
        headers = [
            'Número OP', 'OP Externa', 'Produto', 'Cliente', 
            'Status', 'Prioridade', 'Data Entrada', 'Data Previsão',
            'Preço Unitário', 'Observações'
        ]
        
        # Escrever cabeçalhos
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Buscar OPs com relacionamentos
        ops = OrdemProducao.objects.filter(
            empresa=request.empresa_atual
        ).select_related('cliente', 'produto').order_by('-data_entrada')
        
        # Escrever dados
        for row, op in enumerate(ops, 2):
            try:
                ws.cell(row=row, column=1, value=op.numero_op or '')
                ws.cell(row=row, column=2, value=op.op_externa or '')
                ws.cell(row=row, column=3, value=f"{op.produto.codigo} - {op.produto.referencia}" if op.produto else '')
                ws.cell(row=row, column=4, value=op.cliente.nome if op.cliente else '')
                ws.cell(row=row, column=5, value=op.get_status_display() or '')
                ws.cell(row=row, column=6, value=op.get_prioridade_display() or '')
                ws.cell(row=row, column=7, value=op.data_entrada.strftime('%d/%m/%Y') if op.data_entrada else '')
                ws.cell(row=row, column=8, value=op.data_previsao.strftime('%d/%m/%Y') if op.data_previsao else '')
                ws.cell(row=row, column=9, value=float(op.preco_unitario) if op.preco_unitario else 0)
                ws.cell(row=row, column=10, value=op.observacoes or '')
            except Exception as e:
                # Em caso de erro em uma linha específica, pular para a próxima
                continue
        
        # Ajustar largura das colunas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Criar response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="ops_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        # Em caso de erro geral, retornar uma mensagem de erro
        from django.contrib import messages
        from django.shortcuts import redirect
        messages.error(request, f'Erro ao exportar OPs: {str(e)}')
        return redirect('producao:ops_listar')


@login_required
def materias_prima_listar(request):
    """Lista matérias primas"""
    materias = MateriaPrima.objects.filter(
        empresa=request.empresa_atual
    ).select_related('fornecedor').order_by('nome')
    
    context = {
        'materias': materias,
        'empresa': request.empresa_atual,
    }
    return render(request, 'producao/materias_prima.html', context)


# =============================================================================
# APIs AJAX
# =============================================================================

@login_required
def atualizar_status_op(request, op_id):
    """API para atualizar status da OP"""
    if request.method == 'POST':
        empresa = get_current_empresa()
        op = get_object_or_404(OrdemProducao, id=op_id, empresa=empresa)
        
        novo_status = request.POST.get('status')
        if novo_status in dict(StatusOP.choices):
            op.status = novo_status
            op.save()
            
            return JsonResponse({
                'success': True,
                'status': op.get_status_display()
            })
    
    return JsonResponse({'success': False})


@login_required
def parar_linha(request, linha_id):
    """API para parar/iniciar linha de produção"""
    if request.method == 'POST':
        # Implementar lógica de controle de linha
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False})


@login_required
def buscar_produto(request):
    """API para buscar produto por referência"""
    referencia = request.GET.get('referencia')
    
    if referencia:
        # Implementar busca no cadastro de produtos
        # Por enquanto, retornar dados simulados
        return JsonResponse({
            'success': True,
            'produto': {
                'codigo': f'PROD-{referencia}',
                'descricao': f'Produto encontrado para {referencia}',
                'preco_unitario': '50.00'
            }
        })
    
    return JsonResponse({'success': False})


# =============================================================================
# APIs REST (Django REST Framework)
# =============================================================================

class OrdemProducaoViewSet(TenantMixin, viewsets.ModelViewSet):
    """ViewSet para CRUD completo de Ordens de Produção"""
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status', 'prioridade', 'cliente']
    search_fields = ['numero_op', 'op_externa', 'produto__referencia', 'produto__codigo', 'cliente__nome']
    ordering_fields = ['data_entrada', 'data_previsao', 'prioridade', 'numero_op']
    ordering = ['-data_entrada']
    
    def get_queryset(self):
        return OrdemProducao.objects.filter(
            empresa=self.request.empresa_atual
        ).select_related('cliente', 'produto', 'responsavel')
    
    def get_serializer_class(self):
        if self.action == 'list':
            return OrdemProducaoListSerializer
        elif self.action == 'retrieve':
            return OrdemProducaoDetailSerializer
        else:
            return OrdemProducaoCreateUpdateSerializer
    
    @action(detail=False, methods=['post'])
    def bulk_update_status(self, request):
        """Atualização em lote de status"""
        serializer = BulkUpdateStatusSerializer(data=request.data)
        if serializer.is_valid():
            op_ids = serializer.validated_data['op_ids']
            status_op = serializer.validated_data['status']
            
            updated = OrdemProducao.objects.filter(
                id__in=op_ids,
                empresa=request.empresa_atual
            ).update(status=status_op)
            
            return Response({
                'message': f'{updated} OPs atualizadas com sucesso.',
                'updated_count': updated
            })
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def bulk_update_progress(self, request):
        """Atualização em lote de progresso"""
        serializer = BulkUpdateProgressSerializer(data=request.data)
        if serializer.is_valid():
            updates = serializer.validated_data['updates']
            
            with transaction.atomic():
                for update in updates:
                    try:
                        op = OrdemProducao.objects.get(
                            id=update['op_id'],
                            empresa=request.empresa_atual
                        )
                        op.porcentagem_concluida = update['porcentagem']
                        op.save()
                    except OrdemProducao.DoesNotExist:
                        continue
            
            return Response({
                'message': f'{len(updates)} OPs atualizadas com sucesso.'
            })
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def iniciar_producao(self, request, pk=None):
        """Iniciar produção de uma OP"""
        op = self.get_object()
        if op.status == StatusOP.CADASTRADA:
            op.status = StatusOP.EM_PRODUCAO
            op.data_inicio = date.today()
            op.save()
            return Response({'message': 'Produção iniciada com sucesso.'})
        return Response(
            {'error': 'OP deve estar no status "Cadastrada" para iniciar produção.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    @action(detail=True, methods=['post'])
    def concluir_producao(self, request, pk=None):
        """Concluir produção de uma OP"""
        op = self.get_object()
        if op.status == StatusOP.EM_PRODUCAO:
            op.status = StatusOP.CONCLUIDA
            op.data_conclusao = date.today()
            op.porcentagem_concluida = 100
            op.save()
            return Response({'message': 'Produção concluída com sucesso.'})
        return Response(
            {'error': 'OP deve estar "Em Produção" para ser concluída.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        """Estatísticas para dashboard"""
        empresa = request.empresa_atual
        
        # KPIs básicos
        total_ops = OrdemProducao.objects.filter(empresa=empresa).count()
        ops_em_producao = OrdemProducao.objects.filter(empresa=empresa, status=StatusOP.EM_PRODUCAO).count()
        ops_concluidas = OrdemProducao.objects.filter(empresa=empresa, status=StatusOP.CONCLUIDA).count()
        ops_atrasadas = OrdemProducao.objects.filter(
            empresa=empresa,
            data_previsao__lt=date.today(),
            status__in=[StatusOP.CADASTRADA, StatusOP.EM_PRODUCAO]
        ).count()
        
        # Capacidade
        try:
            capacidade = CapacidadeProducao.objects.get(empresa=empresa)
            capacidade_diaria = capacidade.capacidade_diaria
            utilizacao = (ops_em_producao / capacidade_diaria * 100) if capacidade_diaria > 0 else 0
        except CapacidadeProducao.DoesNotExist:
            capacidade_diaria = 0
            utilizacao = 0
        
        # Valores
        valor_em_producao = OrdemProducao.objects.filter(
            empresa=empresa, status=StatusOP.EM_PRODUCAO
        ).aggregate(total=Sum('preco_unitario'))['total'] or 0
        
        valor_concluido = OrdemProducao.objects.filter(
            empresa=empresa, status=StatusOP.CONCLUIDA
        ).aggregate(total=Sum('preco_unitario'))['total'] or 0
        
        # Matérias primas em baixo estoque
        materias_baixo_estoque = MateriaPrima.objects.filter(
            empresa=empresa,
            estoque_atual__lte=F('estoque_minimo')
        ).count()
        
        # Dados para gráficos
        ops_por_status = {}
        for choice in StatusOP.choices:
            count = OrdemProducao.objects.filter(empresa=empresa, status=choice[0]).count()
            ops_por_status[choice[1]] = count
        
        # Produção por mês (últimos 6 meses)
        hoje = date.today()
        producao_por_mes = []
        for i in range(6):
            mes_ref = hoje - timedelta(days=i*30)
            count = OrdemProducao.objects.filter(
                empresa=empresa,
                data_entrada__year=mes_ref.year,
                data_entrada__month=mes_ref.month
            ).count()
            producao_por_mes.append({
                'mes': mes_ref.strftime('%m/%Y'),
                'count': count
            })
        
        stats_data = {
            'total_ops': total_ops,
            'ops_em_producao': ops_em_producao,
            'ops_concluidas': ops_concluidas,
            'ops_atrasadas': ops_atrasadas,
            'capacidade_diaria': capacidade_diaria,
            'utilizacao_capacidade': round(utilizacao, 2),
            'valor_em_producao': float(valor_em_producao),
            'valor_concluido': float(valor_concluido),
            'materias_primas_baixo_estoque': materias_baixo_estoque,
            'ops_por_status': ops_por_status,
            'producao_por_mes': list(reversed(producao_por_mes)),
            'departamentos_ocupacao': []  # TODO: Implementar após dados reais
        }
        
        serializer = DashboardStatsSerializer(stats_data)
        return Response(serializer.data)


class DepartamentoViewSet(TenantMixin, viewsets.ModelViewSet):
    """ViewSet para CRUD de Departamentos"""
    serializer_class = DepartamentoSerializer
    permission_classes = [IsAuthenticated]
    ordering = ['ordem', 'nome']
    
    def get_queryset(self):
        return Departamento.objects.filter(empresa=self.request.empresa_atual)


class MateriaPrimaViewSet(TenantMixin, viewsets.ModelViewSet):
    """ViewSet para CRUD de Matérias Primas"""
    serializer_class = MateriaPrimaSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['ativo', 'fornecedor']
    search_fields = ['codigo', 'nome', 'descricao']
    ordering = ['nome']
    
    def get_queryset(self):
        return MateriaPrima.objects.filter(
            empresa=self.request.empresa_atual
        ).select_related('fornecedor')
    
    @action(detail=False, methods=['get'])
    def baixo_estoque(self, request):
        """Matérias primas com estoque baixo"""
        materias = self.get_queryset().filter(
            estoque_atual__lte=F('estoque_minimo')
        )
        serializer = self.get_serializer(materias, many=True)
        return Response(serializer.data)


class ProcessoProducaoViewSet(TenantMixin, viewsets.ModelViewSet):
    """ViewSet para controle de processos"""
    serializer_class = ProcessoProducaoSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return ProcessoProducao.objects.filter(
            ordem_producao__empresa=self.request.empresa_atual
        ).select_related('ordem_producao', 'departamento', 'responsavel')
    
    @action(detail=True, methods=['post'])
    def iniciar_processo(self, request, pk=None):
        """Iniciar um processo específico"""
        processo = self.get_object()
        processo.data_inicio = datetime.now()
        processo.save()
        return Response({'message': 'Processo iniciado com sucesso.'})
    
    @action(detail=True, methods=['post'])
    def concluir_processo(self, request, pk=None):
        """Concluir um processo específico"""
        processo = self.get_object()
        processo.data_conclusao = datetime.now()
        processo.porcentagem_concluida = 100
        processo.save()
        return Response({'message': 'Processo concluído com sucesso.'})


# =============================================================================
# AJAX VIEWS
# =============================================================================

@csrf_exempt
@login_required
def atualizar_progresso_op(request):
    """Atualizar progresso de OP via AJAX"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            op_id = data.get('op_id')
            porcentagem = data.get('porcentagem')
            
            op = OrdemProducao.objects.get(
                id=op_id, 
                empresa=request.empresa_atual
            )
            op.porcentagem_concluida = porcentagem
            op.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Progresso atualizado com sucesso.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'success': False}, status=405)


@login_required
def buscar_clientes(request):
    """Busca clientes para autocomplete"""
    query = request.GET.get('q', '')
    clientes = []
    
    if query:
        from apps.cadastros.models import Cliente
        clientes_obj = Cliente.objects.filter(
            empresa=request.empresa_atual,
            nome__icontains=query
        )[:10]
        
        clientes = [
            {'id': c.id, 'text': c.nome}
            for c in clientes_obj
        ]
    
    return JsonResponse({'results': clientes})


@login_required
def buscar_produtos(request):
    """Busca produtos para autocomplete"""
    query = request.GET.get('q', '')
    produtos = []
    
    if query:
        from apps.cadastros.models import Produto
        produtos_obj = Produto.objects.filter(
            empresa=request.empresa_atual,
            referencia__icontains=query
        ).filter(ativo=True)[:10]
        
        produtos = [
            {
                'id': p.id, 
                'text': f"{p.codigo} - {p.referencia}",
                'preco': float(p.preco_unitario) if p.preco_unitario else 0.0
            }
            for p in produtos_obj
        ]
    
    return JsonResponse({'results': produtos})


@login_required
def acompanhar_producao(request):
    """Dashboard de acompanhamento da produção em tempo real"""
    from .models import LinhaProducao, OrdemProducao, ControleEtapaOP, EtapaProducao
    from django.utils import timezone
    
    empresa = request.empresa_atual
    
    # OPs em produção
    ops_em_producao = OrdemProducao.objects.filter(
        empresa=empresa,
        status='EM_PRODUCAO'
    ).select_related('cliente', 'produto', 'linha_producao').prefetch_related(
        'controles_etapa__etapa'
    )
    
    # Preparar dados das OPs com progresso por etapa
    ops_dados = []
    for op in ops_em_producao:
        controles = op.controles_etapa.all().order_by('etapa__ordem')
        
        # Calcular progresso geral
        if controles.exists():
            progresso_total = sum(c.porcentagem_concluida for c in controles) / controles.count()
        else:
            progresso_total = 0
        
        # Identificar etapa atual
        etapa_atual = None
        for controle in controles:
            if controle.status == 'EM_ANDAMENTO':
                etapa_atual = controle.etapa
                break
        
        if not etapa_atual:
            # Se nenhuma em andamento, pegar a próxima não iniciada
            for controle in controles:
                if controle.status == 'NAO_INICIADA':
                    etapa_atual = controle.etapa
                    break
        
        ops_dados.append({
            'op': op,
            'controles': controles,
            'progresso_total': round(progresso_total, 1),
            'etapa_atual': etapa_atual
        })
    
    # Estatísticas do dia
    hoje = timezone.now().date()
    stats_hoje = {
        'ops_iniciadas': OrdemProducao.objects.filter(
            empresa=empresa, data_inicio=hoje
        ).count(),
        'ops_concluidas': OrdemProducao.objects.filter(
            empresa=empresa, data_conclusao=hoje
        ).count(),
        'producao_total': sum(linha.producao_hoje for linha in LinhaProducao.objects.filter(empresa=empresa)),
        'linhas_ativas': LinhaProducao.objects.filter(empresa=empresa, status='ATIVA').count()
    }
    
    # Alertas e problemas
    alertas = []
    
    # OPs atrasadas
    ops_atrasadas = OrdemProducao.objects.filter(
        empresa=empresa,
        status__in=['CADASTRADA', 'EM_PRODUCAO'],
        data_previsao__lt=hoje
    ).count()
    
    if ops_atrasadas > 0:
        alertas.append({
            'tipo': 'danger',
            'icone': 'fas fa-exclamation-triangle',
            'titulo': 'OPs Atrasadas',
            'mensagem': f'{ops_atrasadas} ordem(ns) de produção com prazo vencido'
        })
    
    # Linhas paradas
    linhas_paradas = LinhaProducao.objects.filter(
        empresa=empresa, status__in=['PARADA', 'MANUTENCAO']
    ).count()
    
    if linhas_paradas > 0:
        alertas.append({
            'tipo': 'warning',
            'icone': 'fas fa-tools',
            'titulo': 'Linhas Paradas',
            'mensagem': f'{linhas_paradas} linha(s) de produção fora de operação'
        })
    
    # Etapas com problemas (muito tempo na mesma etapa)
    limite_horas = 8  # 8 horas
    limite_tempo = timezone.now() - timezone.timedelta(hours=limite_horas)
    
    etapas_problemas = ControleEtapaOP.objects.filter(
        ordem_producao__empresa=empresa,
        status='EM_ANDAMENTO',
        data_inicio__lt=limite_tempo
    ).count()
    
    if etapas_problemas > 0:
        alertas.append({
            'tipo': 'info',
            'icone': 'fas fa-clock',
            'titulo': 'Etapas Demoradas',
            'mensagem': f'{etapas_problemas} etapa(s) há mais de {limite_horas}h na mesma fase'
        })
    
    context = {
        'empresa': empresa,
        'ops_dados': ops_dados,
        'stats_hoje': stats_hoje,
        'alertas': alertas
    }
    return render(request, 'producao/acompanhar_producao.html', context)


# APIs AJAX para controle de etapas
@csrf_exempt
@login_required
def api_iniciar_etapa(request, controle_id):
    """API para iniciar uma etapa de produção"""
    if request.method == 'POST':
        try:
            from .models import ControleEtapaOP
            
            controle = get_object_or_404(
                ControleEtapaOP, 
                id=controle_id, 
                ordem_producao__empresa=request.empresa_atual
            )
            
            controle.iniciar_etapa(request.user)
            
            return JsonResponse({
                'success': True,
                'message': f'Etapa {controle.etapa.nome} iniciada com sucesso',
                'status': controle.get_status_display(),
                'status_color': controle.status_color
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'success': False}, status=405)


@csrf_exempt
@login_required
def api_concluir_etapa(request, controle_id):
    """API para concluir uma etapa de produção"""
    if request.method == 'POST':
        try:
            import json
            from .models import ControleEtapaOP
            
            data = json.loads(request.body)
            quantidade_produzida = data.get('quantidade_produzida')
            observacoes = data.get('observacoes', '')
            
            controle = get_object_or_404(
                ControleEtapaOP, 
                id=controle_id, 
                ordem_producao__empresa=request.empresa_atual
            )
            
            # Atualizar observações se fornecidas
            if observacoes:
                controle.observacoes = observacoes
                controle.save()
            
            controle.concluir_etapa(quantidade_produzida, request.user)
            
            return JsonResponse({
                'success': True,
                'message': f'Etapa {controle.etapa.nome} concluída com sucesso',
                'status': controle.get_status_display(),
                'status_color': controle.status_color,
                'porcentagem': float(controle.porcentagem_concluida)
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'success': False}, status=405)


@csrf_exempt
@login_required
def api_pausar_etapa(request, controle_id):
    """API para pausar uma etapa de produção"""
    if request.method == 'POST':
        try:
            from .models import ControleEtapaOP, StatusEtapa
            
            controle = get_object_or_404(
                ControleEtapaOP, 
                id=controle_id, 
                ordem_producao__empresa=request.empresa_atual
            )
            
            if controle.status == StatusEtapa.EM_ANDAMENTO:
                controle.status = StatusEtapa.PAUSADA
                controle.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Etapa {controle.etapa.nome} pausada',
                    'status': controle.get_status_display(),
                    'status_color': controle.status_color
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Etapa deve estar em andamento para ser pausada'
                }, status=400)
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'success': False}, status=405)


@csrf_exempt
@login_required
def api_retomar_etapa(request, controle_id):
    """API para retomar uma etapa pausada"""
    if request.method == 'POST':
        try:
            from .models import ControleEtapaOP, StatusEtapa
            
            controle = get_object_or_404(
                ControleEtapaOP, 
                id=controle_id, 
                ordem_producao__empresa=request.empresa_atual
            )
            
            if controle.status == StatusEtapa.PAUSADA:
                controle.status = StatusEtapa.EM_ANDAMENTO
                controle.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Etapa {controle.etapa.nome} retomada',
                    'status': controle.get_status_display(),
                    'status_color': controle.status_color
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Etapa deve estar pausada para ser retomada'
                }, status=400)
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'success': False}, status=405)


@login_required
def api_status_linha(request, linha_id):
    """API para obter status em tempo real de uma linha"""
    try:
        from .models import LinhaProducao
        
        linha = get_object_or_404(LinhaProducao, id=linha_id, empresa=request.empresa_atual)
        
        return JsonResponse({
            'success': True,
            'linha': {
                'id': linha.id,
                'nome': linha.nome,
                'status': linha.status,
                'status_display': linha.get_status_display(),
                'status_color': linha.status_color,
                'ops_ativas': linha.ops_ativas,
                'producao_hoje': linha.producao_hoje,
                'eficiencia_hoje': round(linha.eficiencia_hoje, 1),
                'capacidade_diaria': linha.capacidade_diaria
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@csrf_exempt
@login_required
def api_alterar_status_linha(request, linha_id):
    """API para alterar status de uma linha de produção"""
    if request.method == 'POST':
        try:
            import json
            from .models import LinhaProducao, StatusLinha
            
            data = json.loads(request.body)
            novo_status = data.get('status')
            
            if novo_status not in [choice[0] for choice in StatusLinha.choices]:
                return JsonResponse({
                    'success': False,
                    'message': 'Status inválido'
                }, status=400)
            
            linha = get_object_or_404(LinhaProducao, id=linha_id, empresa=request.empresa_atual)
            linha.status = novo_status
            linha.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Status da linha alterado para {linha.get_status_display()}',
                'status': linha.status,
                'status_display': linha.get_status_display(),
                'status_color': linha.status_color
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'success': False}, status=405)


@login_required
def avancar_status_op(request, op_id):
    """Avançar status da OP usando workflow"""
    op = get_object_or_404(OrdemProducao, id=op_id, empresa=get_current_empresa())
    
    from .workflow import WorkflowOP
    workflow = WorkflowOP(op)
    proximos_status = workflow.obter_proximos_status(op.status)
    
    if proximos_status and len(proximos_status) > 0:
        novo_status = proximos_status[0]
        op.status = novo_status
        op.save()
        
        messages.success(request, f'Status da OP {op.numero_op} alterado para {op.get_status_display()}')
    else:
        messages.warning(request, f'Não é possível avançar o status da OP {op.numero_op}')
    
    return redirect('producao:ops_listar')


@login_required
def op_excluir(request, op_id):
    """Excluir uma OP"""
    op = get_object_or_404(OrdemProducao, id=op_id, empresa=get_current_empresa())
    
    if request.method == 'POST':
        numero_op = op.numero_op
        try:
            with transaction.atomic():
                # Excluir registros relacionados primeiro
                op.itens_grade.all().delete()
                op.materias_primas.all().delete()
                op.processos.all().delete()
                
                # Excluir controles de etapa se existirem
                if hasattr(op, 'controles_etapa'):
                    op.controles_etapa.all().delete()
                
                # Excluir histórico se existir
                if hasattr(op, 'historicos'):
                    op.historicos.all().delete()
                
                # Excluir a OP
                op.delete()
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': f'OP {numero_op} excluída com sucesso!'
                    })
                else:
                    messages.success(request, f'OP {numero_op} excluída com sucesso!')
                    return redirect('producao:ops_listar')
                    
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Erro ao excluir OP: {str(e)}'
                }, status=400)
            else:
                messages.error(request, f'Erro ao excluir OP: {str(e)}')
                return redirect('producao:ops_listar')
    
    # Se for GET, retornar erro
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': False,
            'message': 'Método não permitido'
        }, status=405)
    else:
        messages.error(request, 'Método não permitido')
        return redirect('producao:ops_listar')


@login_required
def ops_listar_teste(request):
    """Versão simplificada para testar exclusão"""
    empresa = get_current_empresa()
    ops = OrdemProducao.objects.filter(empresa=empresa).select_related('cliente', 'produto')[:10]
    
    context = {
        'ops': ops,
        'empresa': empresa,
    }
    return render(request, 'producao/ops_listar_simples.html', context)
